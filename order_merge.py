"""
壳牌运输订单调度工具
将 SDCC 导出的订单明细表 + 客户邮件发来的订单表，按单号匹配合并，
再按「省份 → 承运商」映射拆分为各承运商的独立 Excel 文件，
并自动发送邮件给各承运商。

用法：把所有 Excel 文件丢入脚本同目录，运行 python3 order_merge.py 即可。
"""

import os
import re
import sys
import smtplib
import socket
from collections import Counter
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import pandas as pd
from base64 import b64encode, b64decode

try:
    import spnego
    HAVE_SPNEGO = True
except ImportError:
    HAVE_SPNEGO = False

# ============================================================
# 配置区
# ============================================================

# 兼容 exe 打包：exe 运行时用 exe 所在目录，脚本运行时用脚本所在目录
if getattr(sys, 'frozen', False):
    SCRIPT_DIR = os.path.dirname(sys.executable)
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

# 承运商 → 省份映射（可通过 .env 覆盖）
def _parse_carrier_config():
    """从 .env 环境变量解析承运商配置，解析失败则用内置默认值"""
    carriers_str = os.environ.get("CARRIER_LIST", "")
    if carriers_str:
        carrier_map = {}
        for name in carriers_str.split("|"):
            name = name.strip()
            if not name:
                continue
            provinces_str = os.environ.get(f"CARRIER_{name}_省", "")
            if provinces_str:
                carrier_map[name] = [p.strip() for p in provinces_str.split(",") if p.strip()]
            else:
                carrier_map[name] = []
        if carrier_map:
            return carrier_map

    # 内置默认值
    return {
        "奥联":   ["黑龙江", "吉林"],
        "富力达": ["河北", "辽宁"],
        "汇利":   ["内蒙古", "新疆", "青海", "河南", "山西", "西藏"],
        "联众":   ["山东"],
        "津京通达": ["天津"],
        "金博通": ["海南", "重庆", "四川", "江西", "福建", "安徽", "湖南", "湖北",
                   "广东", "贵州", "云南", "广西", "浙江", "上海", "江苏"],
    }

CARRIER_MAP = _parse_carrier_config()

# ============================================================
# 从 .env 加载配置
# ============================================================
def _load_env():
    env_path = os.path.join(SCRIPT_DIR, ".env")
    if os.path.exists(env_path):
        with open(env_path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#") and "=" in line:
                    key, val = line.split("=", 1)
                    key, val = key.strip(), val.strip()
                    if key not in os.environ:
                        os.environ[key] = val

_load_env()

# ============================================================
# 邮件配置
# ============================================================
EMAIL_CONFIG = {
    "smtp_server": os.environ.get("EMAIL_SMTP_SERVER", "owa.sinotrans.com"),
    "smtp_port": int(os.environ.get("EMAIL_SMTP_PORT", "587")),
    "sender": os.environ["EMAIL_SENDER"],
    "auth_code": os.environ["EMAIL_PASSWORD"],
    "timeout": 30,
}

# 承运商 → 收件邮箱（可通过 .env 覆盖，多个用分号分隔）
def _parse_carrier_emails():
    """从 .env 读取承运商邮箱，失败则用内置默认值"""
    carriers_str = os.environ.get("CARRIER_LIST", "")
    if carriers_str:
        carrier_emails = {}
        for name in carriers_str.split("|"):
            name = name.strip()
            if not name:
                continue
            emails_str = os.environ.get(f"CARRIER_{name}_邮箱", "")
            if emails_str:
                carrier_emails[name] = emails_str
            else:
                carrier_emails[name] = ""
        if any(v for v in carrier_emails.values()):
            return carrier_emails

    # 内置默认值
    return {
        "奥联":   ";".join([
            "shikun.0101@163.com", "chenjibing@sinotrans.com",
            "yn416@163.com", "shelldispatchtj@sinotrans.com", "15620067562@163.com"
        ]),
        "富力达": ";".join([
            "jinxu_yang@163.com", "chenjibing@sinotrans.com", "shelldispatchtj@sinotrans.com"
        ]),
        "汇利":   ";".join([
            "13821789358@163.com", "chenjibing@sinotrans.com", "shelldispatchtj@sinotrans.com"
        ]),
        "联众":   ";".join([
            "chenjibing@sinotrans.com", "shelldispatchtj@sinotrans.com", "lixuekuan@cmhk.com"
        ]),
        "津京通达": ";".join([
            "shelldispatchtj@sinotrans.com", "zhengdongdong_Jl@163.com"
        ]),
        "金博通": ";".join([
            "tjybwlgs@126.com", "chenjibing@sinotrans.com", "shelldispatchtj@sinotrans.com"
        ]),
    }

CARRIER_EMAILS = _parse_carrier_emails()

# 调度总表收件人 + 未匹配承运商收件人（可通过 .env 覆盖）
MASTER_RECIPIENTS = os.environ.get(
    "MASTER_RECIPIENTS",
    ";".join(["shelldispatchtj@sinotrans.com", "caimeng1@cmhk.com",
              "lixuekuan@cmhk.com", "liushuo2@cmhk.com"])
)
UNMATCHED_RECIPIENT = os.environ.get("UNMATCHED_RECIPIENT", "shelldispatchtj@sinotrans.com")

# 表1（SDCC 导出）列索引（0-based）
T1_ORDER_NO      = 1    # B列 - 单号
T1_ORDER_TYPE    = 20   # U列 - 订单类型
T1_CONSIGNEE     = 36   # AK列 - 收货方
T1_ADDRESS       = 24   # Y列 - 收货地址
T1_CITY          = 25   # Z列 - 城市
T1_GROSS_WEIGHT  = 169  # FN列 - 毛量KG
T1_MATERIAL_NO   = 161  # FF列 - 物料号
T1_MATERIAL_NAME = 162  # FG列 - 物料名称
T1_QUANTITY      = 167  # FL列 - 数量
T1_UNIT_TYPE     = 168  # FM列 - 单位类型(KAR/EA)
T1_VOLUME        = 170  # FO列 - 体积L（需 *1000）
T1_CONTACT       = 50   # AY列 - 收货方联系人
T1_BT_DATE       = 71   # BT列 - SDCC 日期（用于匹配客户子表日期）

# 表2（客户邮件）列索引（0-based）
T2_ORDER_NO  = 0    # A列 - 单号
T2_GKA       = 4    # E列 - GKA
T2_PROVINCE  = 9    # J列 - 省份
T2_REMARK    = 13   # N列 - 备注
T2_OTIF      = 16   # Q列 - OTIF


# ============================================================
# 日期工具（精确到天）
# ============================================================

def parse_date_any(value):
    """尝试解析日期，返回 (year, month, day) 或 (year, month, None) 或 None"""
    if value is None:
        return None
    s = str(value).strip()
    if not s or s.lower() == 'nan':
        return None
    s = re.sub(r'\.0$', '', s)

    # 日期+时间: 2026-06-16 05:02:33
    m = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})\s+\d{1,2}:\d{2}', s)
    if m:
        return (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    # 完整日期: 2026-06-15, 2026/06/15
    m = re.match(r'^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$', s)
    if m:
        return (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    # 中文日期: 2026年6月15日
    m = re.match(r'^(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日', s)
    if m:
        return (int(m.group(1)), int(m.group(2)), int(m.group(3)))
    # 年月: 2026-06, 2026/06
    m = re.match(r'^(\d{4})[-/](\d{1,2})$', s)
    if m:
        return (int(m.group(1)), int(m.group(2)), None)
    # 纯数字: 8位 YYYYMMDD, 6位 YYYYMM
    if s.isdigit():
        if len(s) == 8:
            return (int(s[:4]), int(s[4:6]), int(s[6:8]))
        if len(s) == 6:
            return (int(s[:4]), int(s[4:6]), None)
    return None


def parse_sheet_date(sheet_name, year_hint=None, month_hint=None):
    """解析子表名中的日期，如 '6.15' → (year, month, day)
    year_hint/month_hint 来自文件名上下文"""
    s = str(sheet_name).strip()
    # 0616 (4位 MMDD)
    m = re.match(r'^(\d{2})(\d{2})$', s)
    if m:
        month = int(m.group(1))
        day = int(m.group(2))
        year = year_hint or datetime.now().year
        return (year, month, day)
    # 6.15 或 06.15
    m = re.match(r'^(\d{1,2})\.(\d{1,2})$', s)
    if m:
        month = int(m.group(1))
        day = int(m.group(2))
        year = year_hint or datetime.now().year
        return (year, month, day)
    # 6-15
    m = re.match(r'^(\d{1,2})-(\d{1,2})$', s)
    if m:
        month = int(m.group(1))
        day = int(m.group(2))
        year = year_hint or datetime.now().year
        return (year, month, day)
    # 月份名简写 + day: Jun15, June15
    months_en = {
        'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
        'july':7,'august':8,'september':9,'october':10,'november':11,'december':12,
        'jan':1,'feb':2,'mar':3,'apr':4,'jun':6,'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12
    }
    m = re.match(r'^([a-zA-Z]+)\s*(\d{1,2})$', s)
    if m:
        mon_str = m.group(1).lower()
        if mon_str in months_en:
            year = year_hint or datetime.now().year
            return (year, months_en[mon_str], int(m.group(2)))
    # 回退：用通用解析
    d = parse_date_any(s)
    if d and d[2] is not None:
        return d
    if d and d[2] is None and month_hint:
        return (d[0], d[1], month_hint)
    return None


def extract_filename_date(filename):
    """从客户文件名提取年月上下文，如 '2026 June xxx.xlsx' → (2026, 6, None)"""
    months_en = {
        'january':1,'february':2,'march':3,'april':4,'may':5,'june':6,
        'july':7,'august':8,'september':9,'october':10,'november':11,'december':12,
        'jan':1,'feb':2,'mar':3,'apr':4,'jun':6,'jul':7,'aug':8,'sep':9,'oct':10,'nov':11,'dec':12
    }
    name_lower = filename.lower()
    # 2026 June / 2026_June
    m = re.search(r'(\d{4})\s*[-_]?\s*([a-zA-Z]+)', name_lower)
    if m:
        year = int(m.group(1))
        month_str = m.group(2)
        if month_str in months_en:
            return (year, months_en[month_str], None)
    # 中文: 2026年6月
    m = re.search(r'(\d{4})\s*年\s*(\d{1,2})\s*月', filename)
    if m:
        return (int(m.group(1)), int(m.group(2)), None)
    # 纯数字: 2026-06
    m = re.search(r'(\d{4})[-/_](\d{1,2})', filename)
    if m:
        return (int(m.group(1)), int(m.group(2)), None)
    return None


def date_match(d1, d2):
    """比较两个日期，必须年月日均匹配（None 视为匹配任意天）"""
    if d1 is None or d2 is None:
        return False
    if d1[0] != d2[0] or d1[1] != d2[1]:
        return False
    if d1[2] is not None and d2[2] is not None:
        return d1[2] == d2[2]
    return True


def date_to_str(d):
    """(year, month, day) → 'YYYYMMDD'"""
    if not d:
        return "unknown"
    if d[2] is not None:
        return f"{d[0]}{d[1]:02d}{d[2]:02d}"
    return f"{d[0]}{d[1]:02d}"


# ============================================================
# 输出目录
# ============================================================

def get_output_subdir(base_dir, date_tuple):
    """返回 output/YYYYMMDD/，已存在则递增 YYYYMMDD_第2次/"""
    date_str = date_to_str(date_tuple)
    candidate = os.path.join(base_dir, date_str)
    if not os.path.exists(candidate):
        return candidate
    for i in range(2, 100):
        candidate = os.path.join(base_dir, f"{date_str}_第{i}次")
        if not os.path.exists(candidate):
            return candidate
    return os.path.join(base_dir, f"{date_str}_第99次")


# ============================================================
# 邮件发送
# ============================================================

def _ntlm_login(conn, username, password):
    """NTLM 认证（Exchange 服务器需要）"""
    conn.ehlo("test")
    auth = spnego.client(username, password,
                         hostname=conn.sock.getpeername()[0],
                         service="SMTP", protocol="ntlm")
    code, _ = conn.docmd("AUTH", "NTLM")
    if code != 334:
        raise Exception(f"AUTH NTLM 失败: {code}")
    out_token = auth.step(None)
    code, resp = conn.docmd(b64encode(out_token).decode())
    if code != 334:
        raise Exception(f"NTLM Type1 失败: {code}")
    challenge = b64decode(resp.decode().strip())
    out_token = auth.step(challenge)
    code, _ = conn.docmd(b64encode(out_token).decode())
    if code != 235:
        raise Exception(f"NTLM Type3 失败: {code}")


def send_email(receivers, subject, body, attachment_path):
    """发送带附件的邮件，receivers 为分号分隔的邮箱字符串。成功返回 True"""
    config = EMAIL_CONFIG
    to_list = [r.strip() for r in receivers.split(";") if r.strip()]

    msg = MIMEMultipart()
    msg["From"] = config["sender"]
    msg["To"] = receivers
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))

    with open(attachment_path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment",
                    filename=os.path.basename(attachment_path))
    msg.attach(part)

    def _try_standard():
        for port in [config["smtp_port"], 465, 587]:
            use_ssl = (port == 465)
            try:
                if use_ssl:
                    server = smtplib.SMTP_SSL(config["smtp_server"], port,
                                              timeout=config["timeout"])
                else:
                    server = smtplib.SMTP(config["smtp_server"], port,
                                          timeout=config["timeout"])
                    server.starttls()
                server.login(config["sender"], config["auth_code"])
                server.sendmail(config["sender"], to_list, msg.as_string())
                server.quit()
                return True
            except Exception:
                continue
        return False

    def _try_ntlm():
        if not HAVE_SPNEGO:
            return False
        try:
            server = smtplib.SMTP(config["smtp_server"], config["smtp_port"],
                                  timeout=config["timeout"])
            _ntlm_login(server, os.environ.get("EMAIL_USERNAME", config["sender"]),
                        config["auth_code"])
            server.sendmail(config["sender"], to_list, msg.as_string())
            server.quit()
            return True
        except Exception:
            return False

    if _try_standard():
        return True
    if _try_ntlm():
        return True
    return False


# ============================================================
# 文件读取
# ============================================================

def find_files(directory):
    """扫描目录（含 data/ 子目录），识别 SDCC 文件和客户文件。
    优先使用 data/ 子目录下最新的文件。"""
    all_xlsx = []

    # 1. 扫描主目录
    for f in os.listdir(directory):
        if f.endswith(('.xlsx', '.xls')) and not f.startswith(('~$', '.~')):
            all_xlsx.append(os.path.join(directory, f))

    # 2. 扫描 data/ 子目录（仅最近 7 天）
    data_dir = os.path.join(directory, "data")
    cutoff = datetime.now() - timedelta(days=7)
    if os.path.isdir(data_dir):
        for sub in sorted(os.listdir(data_dir), reverse=True):
            sub_path = os.path.join(data_dir, sub)
            if os.path.isdir(sub_path):
                # 子目录名为 YYYYMMDD 格式时，检查是否在 7 天内
                m = re.match(r'^(\d{4})(\d{2})(\d{2})$', sub)
                if m:
                    sub_date = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                    if sub_date < cutoff:
                        continue
                for f in os.listdir(sub_path):
                    if f.endswith(('.xlsx', '.xls')) and not f.startswith(('~$', '.~')):
                        all_xlsx.append(os.path.join(sub_path, f))

    sdcc_file = None
    customer_files = []

    for f in all_xlsx:
        basename = os.path.basename(f)
        if re.match(r'download_?\d{4}[-_]?\d{2}[-_]?\d{2}', basename, re.IGNORECASE):
            if sdcc_file is None or os.path.getmtime(f) > os.path.getmtime(sdcc_file):
                sdcc_file = f
        else:
            customer_files.append(f)

    return sdcc_file, customer_files


def read_sdcc(filepath):
    """读取 SDCC 导出文件，返回 (clean_df, raw_df)"""
    try:
        df = pd.read_excel(filepath, header=None, dtype=object, engine='openpyxl')
    except Exception as e:
        print(f"  ⚠ 读取 SDCC 文件失败: {e}")
        raise
    df = df.dropna(how='all').reset_index(drop=True)

    cols = {
        'order_no':      T1_ORDER_NO,
        'order_type':    T1_ORDER_TYPE,
        'consignee':     T1_CONSIGNEE,
        'address':       T1_ADDRESS,
        'city':          T1_CITY,
        'gross_weight':  T1_GROSS_WEIGHT,
        'material_no':   T1_MATERIAL_NO,
        'material_name': T1_MATERIAL_NAME,
        'quantity':      T1_QUANTITY,
        'unit_type':     T1_UNIT_TYPE,
        'volume':        T1_VOLUME,
        'contact':       T1_CONTACT,
    }

    result = pd.DataFrame()
    for name, idx in cols.items():
        if idx < len(df.columns):
            result[name] = df.iloc[:, idx]
        else:
            print(f"  ⚠ SDCC 文件列索引 {idx} 超出范围（共 {len(df.columns)} 列），'{name}' 填为空")
            result[name] = None

    # 清洗单号
    result['order_no'] = result['order_no'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)

    # 过滤垃圾行：单号为空、nan、纯空白、不包含数字 的行一律剔除
    def is_valid_order(val):
        s = str(val).strip()
        if not s or s.lower() == 'nan':
            return False
        if not re.search(r'\d', s):
            return False
        return True

    result = result[result['order_no'].apply(is_valid_order)].reset_index(drop=True)
    result['order_no_for_match'] = result['order_no'].str.lstrip('0')
    return result, df


def read_customer(filepath, target_date=None, filename_date=None):
    """读取客户文件的所有 sheet，筛选日期匹配 target_date 的。
    target_date: (year, month, day) 来自 SDCC BT 列
    filename_date: (year, month, None) 来自文件名上下文，用于补齐 sheet 名日期
    返回 list of DataFrame"""
    try:
        xl = pd.ExcelFile(filepath, engine='openpyxl')
    except Exception as e:
        print(f"  ⚠ 读取失败 {os.path.basename(filepath)}: {e}，已跳过")
        return []

    year_hint = filename_date[0] if filename_date else None
    month_hint = filename_date[1] if filename_date else None

    results = []
    for sheet_name in xl.sheet_names:
        sheet_date = parse_sheet_date(sheet_name, year_hint=year_hint, month_hint=month_hint)

        if target_date and sheet_date:
            if not date_match(target_date, sheet_date):
                continue

        try:
            df = pd.read_excel(filepath, sheet_name=sheet_name, header=None, dtype=object, engine='openpyxl')
        except Exception as e:
            print(f"  ⚠ 读取 sheet '{sheet_name}' 失败: {e}，已跳过")
            continue

        df = df.dropna(how='all').reset_index(drop=True)
        if len(df) == 0:
            continue

        cols = {
            'order_no': T2_ORDER_NO,
            'gka':      T2_GKA,
            'province': T2_PROVINCE,
            'remark':   T2_REMARK,
            'otif':     T2_OTIF,
        }
        result = pd.DataFrame()
        for name, idx in cols.items():
            if idx < len(df.columns):
                result[name] = df.iloc[:, idx]
            else:
                result[name] = None

        result['order_no'] = result['order_no'].astype(str).str.strip().str.replace(r'\.0$', '', regex=True)
        result['order_no_for_match'] = result['order_no'].str.strip()
        result['_sheet'] = sheet_name
        results.append(result)

    return results


def extract_sdcc_date(sdcc_raw):
    """从 SDCC 原始 DataFrame 的 BT 列提取完整日期（精确到天）。
    跳过表头行，取出现次数最多的日期。"""
    if T1_BT_DATE >= len(sdcc_raw.columns):
        return None
    col = sdcc_raw.iloc[:, T1_BT_DATE]
    dates = []
    for val in col.dropna():
        d = parse_date_any(val)
        if d and d[2] is not None:
            dates.append((d[0], d[1], d[2]))
    if not dates:
        return None
    most = Counter(dates).most_common(1)[0][0]
    return most


# ============================================================
# 业务逻辑
# ============================================================

def calc_unit(row):
    """根据 FM 列和毛量/数量比值计算单位"""
    unit_type = str(row.get('unit_type', '')).strip().upper()
    if unit_type == 'KAR':
        return '纸箱'
    if unit_type == 'EA':
        try:
            gross = float(row.get('gross_weight', 0) or 0)
            qty = float(row.get('quantity', 0) or 0)
            if qty == 0:
                return 'EA'
            ratio = gross / qty
            if ratio > 500:
                return '大罐'
            elif ratio >= 50:
                return '大桶'
            else:
                return '小桶'
        except (ValueError, TypeError):
            return 'EA'
    return str(row.get('unit_type', ''))


def get_carrier(province):
    """根据省份返回承运商名称"""
    province = str(province).strip()
    for carrier, provinces in CARRIER_MAP.items():
        for p in provinces:
            if p in province or province in p:
                return carrier
    return None


def merge_and_process(sdcc_df, customer_dfs):
    """合并表1和表2，生成总表"""
    if len(customer_dfs) > 0:
        customer = pd.concat(customer_dfs, ignore_index=True)
        customer = customer.drop_duplicates(subset=['order_no_for_match'], keep='first')
    else:
        customer = pd.DataFrame(columns=['order_no_for_match', 'gka', 'province', 'remark', 'otif'])

    merged = sdcc_df.merge(
        customer[['order_no_for_match', 'gka', 'province', 'remark', 'otif']],
        on='order_no_for_match',
        how='left'
    )

    # 省份来自表2 J列，据此匹配承运商
    merged['carrier'] = merged['province'].apply(get_carrier)
    merged['unit_calc'] = merged.apply(calc_unit, axis=1)

    master = pd.DataFrame()
    master['GKA']                  = merged['gka']
    master['订单类型']              = ''
    master['预约时间']              = ''
    master['合提']                  = ''
    master['承运商']                = merged['carrier']
    master['客户订单']              = merged['order_no']
    master['目的地（省）']          = merged['province']
    master['目的地（市）']          = merged['city']
    master['目的地']                = merged['consignee']
    master['目的地（详细地址）']    = merged['address']
    master['收货方联系电话']        = merged['contact']
    master['物料编码']              = merged['material_no']
    master['物料名称']              = merged['material_name']
    master['数量']                  = merged['quantity']
    master['单位']                  = merged['unit_calc']
    master['重量']                  = merged['gross_weight']
    master['体积']                  = pd.to_numeric(merged['volume'], errors='coerce') * 1000
    master['OTIF']                  = merged['otif']
    master['备注']                  = merged['remark']

    master = master.sort_values(
        by=['承运商', '目的地（省）', '目的地（市）'],
        ascending=[True, True, True],
        na_position='last'
    ).reset_index(drop=True)

    return master


# ============================================================
# 输出与发送
# ============================================================

def auto_fit_columns(filepath):
    """打开已生成的 xlsx，自动调整每列宽度"""
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    wb = load_workbook(filepath)
    for ws in wb.worksheets:
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                val = str(cell.value) if cell.value is not None else ''
                char_len = 0
                for ch in val:
                    char_len += 2 if '一' <= ch <= '鿿' or '　' <= ch <= '〿' or '＀' <= ch <= '￯' else 1
                max_len = max(max_len, char_len)
            adjusted = min(max_len + 4, 60)
            adjusted = max(adjusted, 6)
            ws.column_dimensions[col_letter].width = adjusted
    wb.save(filepath)


def write_and_send(master, output_dir):
    """写入总表和各承运商分表，并发送邮件。
    设置环境变量 DRY_RUN=1 可跳过发邮件，仅生成 Excel。"""
    dry_run = os.environ.get("DRY_RUN", "") == "1"
    os.makedirs(output_dir, exist_ok=True)
    date_str = os.path.basename(output_dir.rstrip('/').rstrip('\\'))
    subject = f"壳牌订单调度表_{date_str}"
    body = "请查收今日订单，附件为贵司配送明细。"

    no_carrier = master[master['承运商'].isna() | (master['承运商'] == '')]

    # --- 调度总表（仅「全部订单」sheet）---
    master_path = os.path.join(output_dir, '调度总表.xlsx')
    with pd.ExcelWriter(master_path, engine='openpyxl') as writer:
        master.to_excel(writer, sheet_name='全部订单', index=False)
    auto_fit_columns(master_path)

    print(f"\n✓ 总表: {master_path}")
    print(f"  总行数: {len(master)}")

    # 发送总表
    if dry_run:
        print(f"  [DRY RUN] 跳过发送总表 → {MASTER_RECIPIENTS}")
    else:
        print(f"  发送总表 → {MASTER_RECIPIENTS}", end="")
        if send_email(MASTER_RECIPIENTS, subject, body, master_path):
            print(" ✅")
        else:
            print(" ❌ 发送失败！")

    # --- 未匹配承运商（单独文件）---
    if len(no_carrier) > 0:
        unmatched_path = os.path.join(output_dir, '未匹配承运商.xlsx')
        no_carrier.to_excel(unmatched_path, index=False)
        auto_fit_columns(unmatched_path)
        provinces = no_carrier['目的地（省）'].dropna().unique()
        print(f"  ⚠ 未匹配承运商: {len(no_carrier)} 行 (省份: {list(provinces)}) → {os.path.basename(unmatched_path)}")

        if dry_run:
            print(f"  [DRY RUN] 跳过发送未匹配表 → {UNMATCHED_RECIPIENT}")
        else:
            print(f"  发送未匹配表 → {UNMATCHED_RECIPIENT}", end="")
            if send_email(UNMATCHED_RECIPIENT, subject + "_未匹配", body, unmatched_path):
                print(" ✅")
            else:
                print(" ❌ 发送失败！")

    # --- 各承运商分表 ---
    if dry_run:
        print("\n📁 生成承运商分表（跳过发邮件）...")
    else:
        print("\n📧 发送承运商邮件...")
    for carrier in CARRIER_MAP:
        subset = master[master['承运商'] == carrier]
        if len(subset) == 0:
            continue
        path = os.path.join(output_dir, f'{carrier}.xlsx')
        subset.to_excel(path, index=False)
        auto_fit_columns(path)

        receivers = CARRIER_EMAILS.get(carrier, "")
        if not receivers:
            print(f"  {carrier}: {len(subset)} 行 (未配置邮箱，跳过)")
            continue

        if dry_run:
            print(f"  {carrier}: {len(subset)} 行 → {receivers} [跳过]")
        else:
            print(f"  {carrier}: {len(subset)} 行 → {receivers}", end="")
            if send_email(receivers, subject, body, path):
                print(" ✅")
            else:
                print(" ❌ 发送失败！")

    print("")


# ============================================================
# 主入口
# ============================================================

def main():
    print("=" * 60)
    print("  壳牌运输订单调度工具")
    print("=" * 60)

    sdcc_file, customer_files = find_files(SCRIPT_DIR)

    if not sdcc_file:
        print("\n❌ 未找到 SDCC 导出文件（文件名应包含 download_日期）")
        print(f"   当前目录: {SCRIPT_DIR}")
        return

    print(f"\n📂 SDCC 文件: {sdcc_file}")
    print(f"📂 客户文件: {len(customer_files)} 个")
    for f in customer_files:
        print(f"   - {f}")

    # 读取 SDCC
    print("\n📖 读取 SDCC 文件...")
    sdcc_df, sdcc_raw = read_sdcc(sdcc_file)
    print(f"   SDCC: {len(sdcc_df)} 行, {sdcc_df['order_no'].nunique()} 个单号")

    # 提取 SDCC 日期（BT列，精确到天）
    sdcc_date = extract_sdcc_date(sdcc_raw)
    if sdcc_date:
        print(f"   SDCC 日期 (BT列): {date_to_str(sdcc_date)}")
    else:
        print("   ⚠ 未能从 SDCC BT 列提取日期")

    # 客户子表日期 = SDCC 日期 + 1 天（客户发的是次日发货计划）
    if sdcc_date:
        dt = datetime(sdcc_date[0], sdcc_date[1], sdcc_date[2]) + timedelta(days=1)
        customer_target_date = (dt.year, dt.month, dt.day)
        print(f"   客户子表匹配日期 (SDCC+1): {date_to_str(customer_target_date)}")
    else:
        customer_target_date = None

    # 读取客户文件，按日期匹配子表
    print("\n📖 读取客户文件（按日期匹配子表）...")
    customer_dfs = []
    for f in customer_files:
        fn_date = extract_filename_date(f)
        print(f"   文件 {f}: 文件名日期={date_to_str(fn_date) if fn_date else '无'}")

        sheets = read_customer(f,
            target_date=customer_target_date,
            filename_date=fn_date
        )
        for sdf in sheets:
            customer_dfs.append(sdf)
            print(f"     ✓ 子表 '{sdf['_sheet'].iloc[0]}': {len(sdf)} 行, {sdf['order_no'].nunique()} 个单号")

    if not customer_dfs:
        print("   ⚠ 未读到任何匹配日期的客户数据！")

    # 合并
    print("\n🔄 匹配合并...")
    master = merge_and_process(sdcc_df, customer_dfs)

    matched = (master['GKA'].notna() & (master['GKA'] != '')).sum()
    print(f"   匹配成功: {matched} / {len(master)} 行")

    # 输出目录使用客户发货计划日期（SDCC+1），而非 SDCC 日期
    date_for_output = customer_target_date or sdcc_date
    if not date_for_output:
        for f in customer_files:
            d = extract_filename_date(f)
            if d:
                date_for_output = d
                break
    if not date_for_output:
        date_for_output = (datetime.now().year, datetime.now().month, datetime.now().day)

    output_base = os.path.join(SCRIPT_DIR, 'output')
    output_dir = get_output_subdir(output_base, date_for_output)
    print(f"\n📝 输出目录: {output_dir}")
    write_and_send(master, output_dir)

    print("✅ 完成！")


if __name__ == "__main__":
    main()
