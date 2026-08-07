"""
壳牌订单调度 - 全流程自动化
1. 下载客户邮件附件（今天收到的邮件）
2. 验证客户文件中是否有明天日期的子表（如0808）→ 没有就删文件退出
3. SDCC 导出订单
4. 合并处理 + 发邮件给各承运商

数据存放规则:
- data/{今天}/     ← 邮件附件 + SDCC 导出文件
- output/{明天}/   ← 生成的调度表（明天=子表日期=发货日）
"""

import sys
import os
import re
import glob
import shutil
from datetime import datetime, timedelta

# 兼容 exe 打包
if getattr(sys, 'frozen', False):
	SCRIPT_DIR = os.path.dirname(sys.executable)
else:
	SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

from email_download import download_customer_attachment
from sdcc_export import run_sdcc_export
from order_merge import main as order_merge_main


def _validate_sdcc_data(filepath):
	"""检查 SDCC 导出文件是否有实际订单数据。返回 (has_data, row_count)"""
	try:
		import pandas as pd
		df = pd.read_excel(filepath, header=None, dtype=object, engine='openpyxl')
		df = df.dropna(how='all').reset_index(drop=True)
		if len(df) == 0:
			return False, 0
		# B列(索引1)是单号列，检查是否有包含数字的有效单号
		order_col_idx = 1 if len(df.columns) > 1 else 0
		order_col = df.iloc[:, order_col_idx].astype(str)

		def _valid(val):
			s = val.strip()
			return bool(s and s.lower() != 'nan' and re.search(r'\d', s))

		count = order_col.apply(_valid).sum()
		return count > 0, count
	except Exception as e:
		print(f"  ⚠ 校验 SDCC 文件失败: {e}")
		return False, 0


def find_customer_file():
	"""在 data/{今天}/ 目录下找到客户文件（非 download_ 开头的 xlsx）"""
	today_dir = os.path.join(SCRIPT_DIR, "data", datetime.now().strftime("%Y%m%d"))
	if not os.path.isdir(today_dir):
		return None
	candidates = []
	for f in glob.glob(os.path.join(today_dir, "*.xlsx")):
		basename = os.path.basename(f)
		if not basename.startswith("download_") and not basename.startswith("~$"):
			candidates.append(f)
	if not candidates:
		return None
	return max(candidates, key=os.path.getmtime)


def get_tomorrow_sheet(customer_file):
	"""验证客户文件是否包含明天日期（MMDD，如0808）的子表。
	返回 (month, day)，失败返回 None"""
	tomorrow = datetime.now() + timedelta(days=1)
	expected = tomorrow.strftime("%m%d")  # "0808"

	try:
		from openpyxl import load_workbook
		wb = load_workbook(customer_file, read_only=True)
		sheets = wb.sheetnames
		wb.close()

		if expected in sheets:
			print(f"  客户表子表: {expected}")
			return (tomorrow.month, tomorrow.day)

		# 兼容 "8.8" "08.08" "8-8" 等变体
		mm, dd = tomorrow.month, tomorrow.day
		variants = [f"{mm}.{dd}", f"{mm:02d}.{dd:02d}", f"{mm}-{dd}", f"{mm:02d}-{dd:02d}"]
		for v in variants:
			if v in sheets:
				print(f"  客户表子表: {v}")
				return (tomorrow.month, tomorrow.day)

		print(f"  ⚠ 未找到明天日期子表 '{expected}'")
		print(f"  文件中的子表: {sheets}")
		return None
	except Exception as e:
		print(f"  ⚠ 读取客户文件失败: {e}")
		return None


def calc_sdcc_date(sub_mm, sub_dd):
	"""子表日期 - 1天 = SDCC 导出日期"""
	today = datetime.now()
	sub_date = datetime(today.year, sub_mm, sub_dd)
	if sub_date > today + timedelta(days=60):
		sub_date = datetime(today.year - 1, sub_mm, sub_dd)
	sdcc_date = sub_date - timedelta(days=1)
	return sdcc_date.strftime("%Y-%m-%d")


if __name__ == "__main__":
	print("=" * 60)
	print("  壳牌订单调度 - 全流程自动化")
	print("=" * 60)

	today = datetime.now()
	today_str = today.strftime("%Y%m%d")
	output_base = os.path.join(SCRIPT_DIR, "output")

	# 清理所有 < 今天的旧 output 目录
	if os.path.isdir(output_base):
		for d in os.listdir(output_base):
			m = re.match(r'^(\d{8})', d)
			if m and m.group(1) < today_str:
				p = os.path.join(output_base, d)
				shutil.rmtree(p)
				print(f"🧹 已清理旧输出: {d}")

	# Step 1: 下载客户邮件附件
	print("\n" + "=" * 60)
	print("  Step 1/4: 下载客户邮件附件")
	print("=" * 60)
	customer_downloaded = download_customer_attachment()
	if not customer_downloaded:
		print("\n❌ 今天未收到新的客户发货计划邮件，流程终止。")
		sys.exit(0)

	# Step 2: 验证客户文件子表日期 = 明天
	print("\n" + "=" * 60)
	print("  Step 2/4: 验证客户文件子表日期")
	print("=" * 60)
	customer_file = find_customer_file()
	sdcc_date_str = None
	customer_target_str = None
	if not customer_file:
		print("\n❌ 未找到客户文件，流程终止。")
		sys.exit(0)

	print(f"客户文件: {os.path.basename(customer_file)}")
	md = get_tomorrow_sheet(customer_file)
	if not md:
		print(f"\n❌ 客户文件子表命名异常（未找到明天日期的子表），删除文件并等待下次重试。")
		os.remove(customer_file)
		sys.exit(0)

	sdcc_date_str = calc_sdcc_date(md[0], md[1])
	os.environ["SDCC_DATE"] = sdcc_date_str
	print(f"SDCC 导出日期（子表-1天）: {sdcc_date_str}")

	# 客户目标日期 = 子表日期（即明天）
	sub_date = datetime(today.year, md[0], md[1])
	if sub_date > today + timedelta(days=60):
		sub_date = datetime(today.year - 1, md[0], md[1])
	customer_target_str = sub_date.strftime("%Y%m%d")
	print(f"输出目录: output/{customer_target_str}/")

	# 精确检查：输出目录已存在 → 今天已完成
	if customer_target_str:
		target_output = os.path.join(output_base, customer_target_str)
		if os.path.exists(os.path.join(target_output, "调度总表.xlsx")):
			print(f"\n✅ 今天已完成（{customer_target_str}/调度总表.xlsx 已存在），无需重复执行。")
			sys.exit(0)

	# Step 3: SDCC 导出
	print("\n" + "=" * 60)
	print("  Step 3/4: SDCC 导出订单")
	print("=" * 60)
	sdcc_result = run_sdcc_export(sdcc_date_str)
	if not sdcc_result:
		print("\n❌ SDCC 导出失败，流程终止。")
		sys.exit(1)

	# 校验 SDCC 文件是否有订单数据（壳牌可能尚未推送订单到 SDCC）
	has_data, row_count = _validate_sdcc_data(sdcc_result)
	if not has_data:
		print(f"\n⚠ SDCC 导出文件无订单数据（壳牌可能尚未推送订单到 SDCC）")
		print(f"  删除空文件，等待下次定时任务重试...")
		os.remove(sdcc_result)
		sys.exit(0)
	print(f"✓ SDCC 文件校验通过: {row_count} 行订单数据")

	# Step 4: 合并处理 + 发邮件
	print("\n" + "=" * 60)
	print("  Step 4/4: 订单合并 + 发送邮件")
	print("=" * 60)
	order_merge_main()

	# 清理：只保留本次生成的 output，删除其他所有旧 output
	if os.path.isdir(output_base) and customer_target_str:
		for d in os.listdir(output_base):
			m = re.match(r'^(\d{8})', d)
			if m and m.group(1) != customer_target_str:
				p = os.path.join(output_base, d)
				shutil.rmtree(p)
				print(f"🧹 已清理旧输出: {d}")

	# 清理 7 天前的 data/ 子目录
	data_dir = os.path.join(SCRIPT_DIR, "data")
	if os.path.isdir(data_dir):
		cutoff = today - timedelta(days=7)
		for sub in os.listdir(data_dir):
			sub_path = os.path.join(data_dir, sub)
			if os.path.isdir(sub_path):
				m = re.match(r'^(\d{4})(\d{2})(\d{2})$', sub)
				if m:
					sub_date = datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
					if sub_date < cutoff:
						shutil.rmtree(sub_path)
						print(f"🧹 已清理旧数据: data/{sub}")

	print("\n" + "=" * 60)
	print("  ✅ 全流程完成！")
	print("=" * 60)
